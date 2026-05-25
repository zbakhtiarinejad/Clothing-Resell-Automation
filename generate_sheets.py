import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_automated_system():
    wb = openpyxl.Workbook()
    
    tabs = [
        "Dashboard", "Inventory", "Sales Log", "Fees & Postage Table",
        "Suppliers & COGS Tracking", "SKU Generator", "Category Performance", "Sell Through Engine"
    ]
    
    for tab in tabs:
        wb.create_sheet(title=tab)
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
        
    font_family = "Segoe UI"
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    title_font = Font(name=font_family, size=16, bold=True, color="1F4E79")
    kpi_lbl_font = Font(name=font_family, size=9, bold=False, color="595959")
    kpi_val_font = Font(name=font_family, size=18, bold=True, color="000000")
    kpi_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    def style_sheet_headers(ws, headers):
        ws.append(headers)
        ws.row_dimensions[1].height = 26
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws_fees = wb["Fees & Postage Table"]
    style_sheet_headers(ws_fees, ["Platform", "Fixed Fee", "Variable Rate", "Avg Postage"])
    ws_fees.append(["eBay", 0.30, 0.128, 4.50])
    ws_fees.append(["Vinted", 0.00, 0.00, 0.00]) 
    ws_fees.append(["Depop", 0.00, 0.10, 4.00])

    ws_inv = wb["Inventory"]
    style_sheet_headers(ws_inv, ["SKU", "Brand", "Category", "Size", "COGS", "Supplier", "Status", "Date Listed"])
    ws_inv.views.sheetView[0].showGridLines = True

    ws_sales = wb["Sales Log"]
    style_sheet_headers(ws_sales, [
        "SKU", "Sale Price", "Platform", "Date Sold", 
        "COGS (Auto)", "Platform Fee (Auto)", "Postage (Auto)", "Net Profit (Auto)", "Days to Sell (Auto)"
    ])
    
    for r in range(2, 1500):
        row = str(r)
        ws_sales[f"E{row}"] = f"=IF(ISBLANK(A{row}), \"\", XLOOKUP(A{row}, Inventory!A:A, Inventory!E:E, 0))"
        ws_sales[f"F{row}"] = f"=IF(ISBLANK(A{row}), \"\", (B{row} * XLOOKUP(C{row}, 'Fees & Postage Table'!A:A, 'Fees & Postage Table'!C:C, 0)) + XLOOKUP(C{row}, 'Fees & Postage Table'!A:A, 'Fees & Postage Table'!B:B, 0))"
        ws_sales[f"G{row}"] = f"=IF(ISBLANK(A{row}), \"\", XLOOKUP(C{row}, 'Fees & Postage Table'!A:A, 'Fees & Postage Table'!D:D, 0))"
        ws_sales[f"H{row}"] = f"=IF(ISBLANK(A{row}), \"\", B{row} - E{row} - F{row} - G{row})"
        ws_sales[f"I{row}"] = f"=IF(ISBLANK(A{row}), \"\", D{row} - XLOOKUP(A{row}, Inventory!A:A, Inventory!H:H, TODAY()))"

    ws_cat = wb["Category Performance"]
    style_sheet_headers(ws_cat, ["Category", "Total Revenue", "Total Net Profit", "Items Sold"])
    ws_cat["A2"] = "=UNIQUE(FILTER(Inventory!C2:C2000, Inventory!C2:C2000<>\"\"))"
    ws_cat["B2"] = "=SUMIFS('Sales Log'!B:B, 'Sales Log'!A:A, A2#)"
    ws_cat["C2"] = "=SUMIFS('Sales Log'!H:H, 'Sales Log'!A:A, A2#)"
    ws_cat["D2"] = "=COUNTIFS('Sales Log'!A:A, A2#)"

    ws_ste = wb["Sell Through Engine"]
    style_sheet_headers(ws_ste, ["Category", "Total Listed", "Total Sold", "Sell-Through Rate"])
    ws_ste["A2"] = "=UNIQUE(FILTER(Inventory!C2:C2000, Inventory!C2:C2000<>\"\"))"
    ws_ste["B2"] = "=COUNTIFS(Inventory!C:C, A2#)"
    ws_ste["C2"] = "=COUNTIFS('Sales Log'!A:A, XLOOKUP(A2#, Inventory!C:C, Inventory!A:A, \"\"))"
    ws_ste["D2"] = "=IF(B2#=0, 0, C2# / B2#)"

    ws_sku = wb["SKU Generator"]
    style_sheet_headers(ws_sku, ["Brand Initials", "Category Code", "Sequence", "Generated SKU"])
    ws_sku.append(["NK", "TSH", 1001, '=CONCATENATE(A2, "-", B2, "-", C2)'])

    ws_sup = wb["Suppliers & COGS Tracking"]
    style_sheet_headers(ws_sup, ["Supplier", "Total Invested", "Total Returns", "Supplier ROI"])
    ws_sup["A2"] = "=UNIQUE(FILTER(Inventory!F2:F2000, Inventory!F2:F2000<>\"\"))"
    ws_sup["B2"] = "=SUMIFS(Inventory!E:E, Inventory!F:F, A2#)"
    ws_sup["C2"] = "=SUMIFS('Sales Log'!H:H, 'Sales Log'!A:A, XLOOKUP(A2#, Inventory!F:F, Inventory!A:A, \"\"))"
    ws_sup["D2"] = "=IF(B2#=0, 0, C2# / B2#)"

    ws_dash = wb["Dashboard"]
    ws_dash.views.sheetView[0].showGridLines = True
    ws_dash["A1"] = "RESALE BUSINESS MANAGEMENT HUBS"
    ws_dash["A1"].font = title_font
    
    kpis = [
        ("B3", "B4", "=SUM('Sales Log'!B:B)", "TOTAL REVENUE"),
        ("D3", "D4", "=SUM('Sales Log'!H:H)", "NET PROFIT"),
        ("F3", "F4", "=SUMIFS(Inventory!E:E, Inventory!G:G, \"Active\")", "CASH TIED IN STOCK"),
        ("H3", "H4", "=AVERAGE('Sales Log'!I:I)", "AVG DAYS TO SELL"),
        ("J3", "J4", "=COUNTIFS(Inventory!G:G, \"Active\")", "ACTIVE INVENTORY VALUE")
    ]
    
    for lbl_c, val_c, formula, title in kpis:
        ws_dash[lbl_c] = title
        ws_dash[lbl_c].font = kpi_lbl_font
        ws_dash[lbl_c].fill = kpi_fill
        ws_dash[lbl_c].alignment = Alignment(horizontal="center", vertical="center")
        ws_dash[val_c] = formula
        ws_dash[val_c].font = kpi_val_font
        ws_dash[val_c].fill = kpi_fill
        ws_dash[val_c].alignment = Alignment(horizontal="center", vertical="center")
        
    ws_dash["B7"] = "Top Performance Categories"
    ws_dash["B7"].font = Font(name=font_family, size=12, bold=True)
    ws_dash["B8"] = "=SORT('Category Performance'!A2:D20, 3, -1)" 

    ws_dash["F7"] = "Supplier ROI Standings"
    ws_dash["F7"].font = Font(name=font_family, size=12, bold=True)
    ws_dash["F8"] = "=SORT('Suppliers & COGS Tracking'!A2:D20, 4, -1)" 

    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max(max_len + 3, 14)

    wb.save("client_delivery/Fully_Automated_System.xlsx")
    print("SUCCESS: Relational spreadsheet built inside client_delivery/ folder.")

if __name__ == "__main__":
    create_automated_system()